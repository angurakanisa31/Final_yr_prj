import io
import csv
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import datetime

from backend.database import get_db
from backend.models import Customer, Product, Purchase, Review, User

router = APIRouter(prefix="/api/reports", tags=["Analytics Reporting"])

@router.get("/customer/csv")
def export_customer_csv(db: Session = Depends(get_db)):
    """Export Customer report in CSV format."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(["Customer ID", "Name", "Segment", "Total Spending ($)", "Purchases Count", "Churn Score (%)", "Lifetime Value"])
    
    customers = db.query(Customer).all()
    for c in customers:
        purchases = db.query(Purchase).filter(Purchase.customer_id == c.id).all()
        total_spending = sum(p.price * p.quantity for p in purchases)
        purchases_count = len(purchases)
        
        writer.writerow([
            c.id,
            c.name,
            c.segment,
            f"{total_spending:.2f}",
            purchases_count,
            f"{c.churn_risk * 100:.1f}",
            c.clv_value
        ])
        
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=customer_report.csv"}
    )


@router.get("/customer/excel")
def export_customer_excel(db: Session = Depends(get_db)):
    """Export Customer report in Excel format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Analysis"
    
    # Styling
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    headers = ["Customer ID", "Name", "Segment", "Total Spending ($)", "Purchases Count", "Churn Score (%)", "Lifetime Value"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    customers = db.query(Customer).all()
    for c in customers:
        purchases = db.query(Purchase).filter(Purchase.customer_id == c.id).all()
        total_spending = sum(p.price * p.quantity for p in purchases)
        purchases_count = len(purchases)
        
        ws.append([
            c.id,
            c.name,
            c.segment,
            round(total_spending, 2),
            purchases_count,
            round(c.churn_risk * 100, 1),
            c.clv_value
        ])
        
    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customer_report.xlsx"}
    )


@router.get("/customer/pdf")
def export_customer_pdf(db: Session = Depends(get_db)):
    """Export Customer report in PDF format."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1F497D'),
        spaceAfter=15
    )
    
    story = []
    
    # Header
    story.append(Paragraph("E-Commerce CRM - Customer Executive Report", title_style))
    story.append(Paragraph(f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 15))
    
    # Table data
    table_data = [["Cust ID", "Name", "Segment", "Total Spend ($)", "Purchases", "Churn Risk", "CLV"]]
    
    customers = db.query(Customer).all()
    for c in customers:
        purchases = db.query(Purchase).filter(Purchase.customer_id == c.id).all()
        total_spending = sum(p.price * p.quantity for p in purchases)
        
        table_data.append([
            str(c.id),
            c.name,
            c.segment,
            f"${total_spending:.2f}",
            str(len(purchases)),
            f"{c.churn_risk * 100:.1f}%",
            c.clv_value
        ])
        
    t = Table(table_data, colWidths=[50, 110, 110, 80, 60, 70, 70])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F497D')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F2F2F2')),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#D3D3D3')),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    
    story.append(t)
    doc.build(story)
    
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=customer_report.pdf"}
    )


@router.get("/company/excel")
def export_company_excel(db: Session = Depends(get_db)):
    """Export Company report in Excel format."""
    wb = Workbook()
    
    # Sheet 1: Sales Summary
    ws_sales = wb.active
    ws_sales.title = "Sales Analytics"
    
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    ws_sales.append(["Metric", "Value"])
    ws_sales.append(["Total Orders", db.query(Purchase).count()])
    total_rev = sum(p.price * p.quantity for p in db.query(Purchase).all())
    ws_sales.append(["Total Revenue ($)", round(total_rev, 2)])
    
    # Retention and recommendations accuracy
    total_customers = db.query(Customer).count()
    high_risk_count = db.query(Customer).filter(Customer.churn_risk >= 0.8).count()
    retention_rate = ((total_customers - high_risk_count) / total_customers * 100) if total_customers > 0 else 100.0
    ws_sales.append(["Customer Retention Rate (%)", round(retention_rate, 1)])
    
    for cell in ws_sales[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    # Sheet 2: Products Performance
    ws_prods = wb.create_sheet(title="Product Performance")
    ws_prods.append(["Product ID", "Product Name", "Brand", "Category", "Price ($)", "Stock Left", "Quantity Sold"])
    for cell in ws_prods[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    products = db.query(Product).all()
    for p in products:
        sales_q = db.query(func.sum(Purchase.quantity)).filter(Purchase.product_id == p.id).scalar() or 0
        ws_prods.append([p.id, p.name, p.brand_name, p.category, p.price, p.stock, sales_q])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=company_report.xlsx"}
    )


@router.get("/company/pdf")
def export_company_pdf(db: Session = Depends(get_db)):
    """Export Company report in PDF format."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#2E7D32'),
        spaceAfter=15
    )
    
    story = []
    story.append(Paragraph("E-Commerce CRM - Company Analytics Report", title_style))
    story.append(Spacer(1, 15))
    
    # Financial metrics table
    purchases = db.query(Purchase).all()
    total_rev = sum(p.price * p.quantity for p in purchases)
    total_customers = db.query(Customer).count()
    high_risk_count = db.query(Customer).filter(Customer.churn_risk >= 0.8).count()
    retention_rate = ((total_customers - high_risk_count) / total_customers * 100) if total_customers > 0 else 100.0
    
    summary_data = [
        ["Key Business Metric", "Value"],
        ["Total Platform Revenue", f"${total_rev:.2f}"],
        ["Total Registered Customers", str(total_customers)],
        ["Customer Retention Rate", f"{retention_rate:.1f}%"],
        ["Total Orders Processed", str(len(purchases))]
    ]
    t_summary = Table(summary_data, colWidths=[200, 150])
    t_summary.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#CCCCCC')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#FFFFFF'), colors.HexColor('#F9F9F9')]),
        ('FONTSIZE', (0,0), (-1,-1), 10),
    ]))
    
    story.append(Paragraph("1. Executive Summary", styles['Heading2']))
    story.append(Spacer(1, 5))
    story.append(t_summary)
    story.append(Spacer(1, 20))
    
    # Product distribution table
    story.append(Paragraph("2. Top Product Analytics", styles['Heading2']))
    story.append(Spacer(1, 5))
    
    prod_data = [["Product", "Brand", "Category", "Price", "Units Sold"]]
    products = db.query(Product).all()
    for p in products[:5]:  # Top products
        sales_q = db.query(func.sum(Purchase.quantity)).filter(Purchase.product_id == p.id).scalar() or 0
        prod_data.append([p.name, p.brand_name, p.category, f"${p.price:.2f}", str(sales_q)])
        
    t_prod = Table(prod_data, colWidths=[150, 100, 100, 80, 80])
    t_prod.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#F5F5F5'), colors.HexColor('#FFFFFF')]),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    story.append(t_prod)
    
    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=company_report.pdf"}
    )
